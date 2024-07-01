#################################
# List all non-standard packages to be imported by your 
# script here (only missing packages will be installed)
from ayx import Package
#Package.installPackages(['pandas','numpy'])
import importlib.util
import sys

name = 'openpyxl'

if name in sys.modules:
    print(f"{name!r} already in sys.modules")
elif (spec := importlib.util.find_spec(name)) is not None:
    # If you chose to perform the actual import ...
    module = importlib.util.module_from_spec(spec)
    sys.modules[name] = module
    spec.loader.exec_module(module)
    print(f"{name!r} has been imported")
else:
    try:
        Package.installPackages(name)
        module = importlib.util.module_from_spec(spec)
        sys.modules[name] = module
        spec.loader.exec_module(module)
        print(f"{name!r} has been imported")
    except:
        print(f"can't find or install the {name!r} module")



#################################
from ayx import Alteryx

AlteryxDF=Alteryx.read("#1")

import importlib.util
import sys
import pandas as pd
try:
    from openpyxl import load_workbook
except:
    print('Modules missing')


# For illustrative purposes.

def sheet_copier(df):
    listOfUniques = df['Full Path'].unique()
    for unique in listOfUniques:
        wb = load_workbook(unique)
        tempdf = df[ df['Full Path'] == unique]
        OrgSheet=tempdf['Original Sheet'].unique()
        for sheet in OrgSheet:
            newSheet=tempdf[tempdf['Original Sheet']==sheet]
            listofNewSheets=newSheet['New Sheet'].unique()
            for newCopy in listofNewSheets:
                dup_ws = wb.copy_worksheet(wb[sheet])
                dup_ws.title = newCopy
                wb.create_sheet(newCopy)
                if f'{newCopy}1' in wb.sheetnames:
                    wb.remove(wb[f'{newCopy}1'])
                print(f'{newCopy} has been created in {unique} from {sheet}')
                wb.save(unique)
                
name = 'openpyxl'

if name in sys.modules:
    sheet_copier(AlteryxDF)
else:
    print(f"can't find the {name!r} module")




#################################
